# -*- coding: utf-8 -*-
"""엑셀 / 마크다운 산출물 생성."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

from judge import APPLIED, NOT_APPLIED, CHECK, NA

NAVY = "1F3864"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
SEC_FILL = PatternFill("solid", fgColor="D9E2F3")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
NG_FILL = PatternFill("solid", fgColor="FFC7CE")
CK_FILL = PatternFill("solid", fgColor="FFEB9C")
NA_FILL = PatternFill("solid", fgColor="E7E6E6")

THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY = Font(name="맑은 고딕", size=10)
SMALL = Font(name="맑은 고딕", size=9)
MONO = Font(name="D2Coding", size=9)
SEC_FONT = Font(name="맑은 고딕", size=10, bold=True, color=NAVY)
WRAP = Alignment(wrap_text=True, vertical="center")
TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

FILLS = {APPLIED: OK_FILL, NOT_APPLIED: NG_FILL, CHECK: CK_FILL, NA: NA_FILL}


def _hdr(ws, row, col0, names, height=22):
    for i, h in enumerate(names, start=col0):
        c = ws.cell(row, i, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = height


# =============================================================================
# 1) 적용 보안설정 목록  (안내용)
# =============================================================================
def build_settings_xlsx(profile, outpath):
    wb = Workbook()
    ws = wb.active
    ws.title = "적용 보안설정"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([3, 20, 6, 34, 56, 32], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws["B2"] = "서버 표준화 스크립트 적용 보안설정"
    ws["B2"].font = Font(name="맑은 고딕", size=16, bold=True, color=NAVY)
    ws["B3"] = "%s  |  대상: %s  |  실행: %s" % (profile["script"], profile["title"], profile["usage"])
    ws["B3"].font = Font(name="맑은 고딕", size=10, color="595959")
    ws["B4"] = "※ 신규 서버 구축 시 본 스크립트를 1회 실행하여 아래 설정을 일괄 적용합니다."
    ws["B4"].font = Font(name="맑은 고딕", size=10, color="595959")

    r = 6
    _hdr(ws, r, 2, ["구분", "No", "보안 설정 항목", "적용 내용", "설정 위치"], 24)
    r += 1
    start = r
    cur, sec_start, merges = None, None, []
    for i, (sec, item, val, loc) in enumerate(profile["settings"], start=1):
        if sec != cur:
            if cur is not None:
                merges.append((sec_start, r - 1))
            cur, sec_start = sec, r
        for col, v in [(2, sec), (3, i), (4, item), (5, val), (6, loc)]:
            c = ws.cell(r, col, v)
            c.border = BORDER
            if col == 2:
                c.font, c.fill, c.alignment = SEC_FONT, SEC_FILL, CENTER
            elif col == 3:
                c.font, c.alignment = BODY, CENTER
            elif col == 6:
                c.font, c.alignment = MONO, WRAP
            else:
                c.font, c.alignment = BODY, WRAP
        ws.row_dimensions[r].height = 30 if ("\n" in val or "\n" in loc) else 21
        r += 1
    merges.append((sec_start, r - 1))
    for a, b in merges:
        if b > a:
            ws.merge_cells(start_row=a, start_column=2, end_row=b, end_column=2)
    ws.freeze_panes = "B%d" % start

    r += 2
    ws.cell(r, 2, "스크립트 실행 후 수동 조치 사항").font = Font(name="맑은 고딕", size=12, bold=True, color=NAVY)
    r += 1
    _hdr(ws, r, 2, ["구분", "항목", "명령", "비고"])
    r += 1
    for item, cmd, note in profile["manual"]:
        for col, v in [(2, "수동"), (3, item), (4, cmd), (5, note)]:
            c = ws.cell(r, col, v)
            c.border, c.fill = BORDER, NOTE_FILL
            c.font = MONO if col == 4 else BODY
            c.alignment = CENTER if col == 2 else WRAP
        ws.row_dimensions[r].height = 30 if "\n" in note else 22
        r += 1

    wb.save(outpath)
    return outpath


def build_settings_md(profile, outpath):
    L = []
    L.append("# 서버 표준화 스크립트 적용 보안설정 — %s" % profile["title"])
    L.append("")
    L.append("| | |")
    L.append("|---|---|")
    L.append("| 스크립트 | `%s` |" % profile["script"])
    L.append("| 대상 OS | %s |" % profile["title"])
    L.append("| 실행 방법 | `%s` |" % profile["usage"])
    L.append("| 설정 항목 수 | %d개 |" % len(profile["settings"]))
    L.append("")
    L.append("신규 서버 구축 시 본 스크립트를 1회 실행하여 아래 설정을 일괄 적용합니다.")
    L.append("")

    cur = None
    for sec, item, val, loc in profile["settings"]:
        if sec != cur:
            cur = sec
            L.append("")
            L.append("## %s" % sec)
            L.append("")
            L.append("| 보안 설정 항목 | 적용 내용 | 설정 위치 |")
            L.append("|---|---|---|")
        L.append("| %s | %s | %s |" % (item, val.replace("\n", "<br>"),
                                       "`%s`" % loc.replace("\n", "`<br>`") if loc != "-" else "-"))

    L.append("")
    L.append("## 스크립트 실행 후 수동 조치 사항")
    L.append("")
    L.append("| 항목 | 명령 | 비고 |")
    L.append("|---|---|---|")
    for item, cmd, note in profile["manual"]:
        L.append("| %s | `%s` | %s |" % (item, cmd, note.replace("\n", "<br>")))
    L.append("")
    L.append("---")
    L.append("")
    L.append("적용 여부 확인은 `verify/verify.py` 로 자동 점검할 수 있습니다.")
    L.append("")
    L.append("```bash")
    L.append("python3 verify/verify.py --profile %s --host <서버IP> --user <계정>" %
             [k for k, v in __import__("profiles").PROFILES.items() if v is profile][0])
    L.append("```")
    L.append("")

    with open(outpath, "w") as f:
        f.write("\n".join(L))
    return outpath


# =============================================================================
# 2) 점검 결과
# =============================================================================
def build_result_xlsx(profile, results, meta, outpath):
    checks = profile["checks"]
    wb = Workbook()

    # ---- 표지 ----
    ws = wb.active
    ws.title = "표지"
    ws.sheet_view.showGridLines = False
    for col, w in [("A", 4), ("B", 24), ("C", 62)]:
        ws.column_dimensions[col].width = w

    ws["B2"] = "서버 표준화 스크립트 적용 점검 결과"
    ws["B2"].font = Font(name="맑은 고딕", size=17, bold=True, color=NAVY)
    ws.merge_cells("B2:C2")
    ws["B3"] = "%s 적용 여부 확인" % profile["script"]
    ws["B3"].font = Font(name="맑은 고딕", size=10, color="595959")
    ws.merge_cells("B3:C3")

    info = [
        ("점검 대상 서버", meta.get("ip", "")),
        ("호스트명", meta.get("host", "")),
        ("운영체제", meta.get("os", "")),
        ("적용 스크립트", profile["script"]),
        ("점검 일시", meta.get("date", "")),
        ("관리 대상 계정", meta.get("tu", "")),
        ("점검자", meta.get("auditor", "")),
        ("총 점검 항목", "%d 건" % len(checks)),
    ]
    r = 5
    for k, v in info:
        a = ws.cell(r, 2, k)
        a.font = Font(name="맑은 고딕", size=10, bold=True)
        a.fill = PatternFill("solid", fgColor="F2F2F2")
        a.border, a.alignment = BORDER, WRAP
        b = ws.cell(r, 3, v)
        b.font, b.border, b.alignment = BODY, BORDER, WRAP
        ws.row_dimensions[r].height = 20
        r += 1

    r += 1
    ws.cell(r, 2, "판정 기준").font = Font(name="맑은 고딕", size=11, bold=True, color=NAVY)
    r += 1
    for k, v in [(APPLIED, "스크립트 설정이 정상 적용되어 있음"),
                 (NOT_APPLIED, "설정이 적용되지 않았거나 이후 변경됨 — 조치 필요"),
                 (CHECK, "자동 판정 불가 — 담당자 확인 필요"),
                 (NA, "해당 없음 (미설치 / 미해당 환경)")]:
        a = ws.cell(r, 2, k)
        a.font = Font(name="맑은 고딕", size=10, bold=True)
        a.fill, a.border, a.alignment = FILLS[k], BORDER, CENTER
        b = ws.cell(r, 3, v)
        b.font, b.border, b.alignment = BODY, BORDER, WRAP
        ws.row_dimensions[r].height = 18
        r += 1

    # ---- 요약 ----
    ws2 = wb.create_sheet("점검 요약")
    ws2.sheet_view.showGridLines = False
    for i, w in enumerate([4, 24, 10, 10, 10, 10, 10, 40], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2["B2"] = "구분별 적용 현황"
    ws2["B2"].font = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)
    _hdr(ws2, 4, 2, ["구분", "항목 수", APPLIED, NOT_APPLIED, CHECK, NA, "조치 대상"])

    secs = []
    for c in checks:
        if c["sec"] not in secs:
            secs.append(c["sec"])

    row, tot = 5, {APPLIED: 0, NOT_APPLIED: 0, CHECK: 0, NA: 0}
    for s in secs:
        items = [c for c in checks if c["sec"] == s]
        cnt = {APPLIED: 0, NOT_APPLIED: 0, CHECK: 0, NA: 0}
        ng = []
        for it in items:
            v = results.get(it["no"], {}).get("verdict", "")
            if v in cnt:
                cnt[v] += 1
                tot[v] += 1
            if v == NOT_APPLIED:
                ng.append(it["no"])
        vals = [s, len(items), cnt[APPLIED], cnt[NOT_APPLIED], cnt[CHECK], cnt[NA], ", ".join(ng)]
        for i, v in enumerate(vals, start=2):
            c = ws2.cell(row, i, v)
            c.font, c.border = BODY, BORDER
            c.alignment = WRAP if i in (2, 8) else CENTER
        if cnt[NOT_APPLIED]:
            ws2.cell(row, 5).fill = NG_FILL
        ws2.row_dimensions[row].height = 20
        row += 1
    for i, v in enumerate(["합 계", len(checks), tot[APPLIED], tot[NOT_APPLIED], tot[CHECK], tot[NA], ""], start=2):
        c = ws2.cell(row, i, v)
        c.font = Font(name="맑은 고딕", size=10, bold=True)
        c.fill, c.border = SEC_FILL, BORDER
        c.alignment = WRAP if i == 2 else CENTER
    ws2.row_dimensions[row].height = 22

    # ---- 상세 ----
    ws3 = wb.create_sheet("점검 결과")
    ws3.sheet_view.showGridLines = False
    cols = ["구분", "No", "점검 항목", "기대값", "판정", "실제 확인 결과", "비고"]
    for i, w in enumerate([20, 7, 32, 34, 10, 62, 30], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3["A1"] = "점검 결과 — %s (%s)" % (meta.get("ip", ""), meta.get("host", ""))
    ws3["A1"].font = Font(name="맑은 고딕", size=13, bold=True, color=NAVY)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws3.row_dimensions[1].height = 26
    _hdr(ws3, 2, 1, cols, 24)

    row, cur, sec_start, merges = 3, None, None, []
    for it in checks:
        if it["sec"] != cur:
            if cur is not None:
                merges.append((sec_start, row - 1))
            cur, sec_start = it["sec"], row
        res = results.get(it["no"], {})
        v = res.get("verdict", "")
        actual = res.get("out", "")
        if len(actual) > 1000:
            actual = actual[:1000] + "\n…(생략)"
        for i, val in enumerate([it["sec"], it["no"], it["item"], it["exp"], v, actual, res.get("note", "")], start=1):
            c = ws3.cell(row, i, val)
            c.border = BORDER
            if i == 1:
                c.font, c.fill, c.alignment = SEC_FONT, SEC_FILL, CENTER
            elif i in (2, 5):
                c.font = Font(name="맑은 고딕", size=9, bold=(i == 5))
                c.alignment = CENTER
            elif i == 6:
                c.font, c.alignment = MONO, TOP
            else:
                c.font, c.alignment = SMALL, TOP
        if v in FILLS:
            ws3.cell(row, 5).fill = FILLS[v]
        ws3.row_dimensions[row].height = min(140, max(28, 13 * min(actual.count("\n") + 1, 9)))
        row += 1
    merges.append((sec_start, row - 1))
    for a, b in merges:
        if b > a:
            ws3.merge_cells(start_row=a, start_column=1, end_row=b, end_column=1)
    last = row - 1

    dv = DataValidation(type="list", formula1='"%s,%s,%s,%s"' % (APPLIED, NOT_APPLIED, CHECK, NA),
                        allow_blank=True, showDropDown=False)
    ws3.add_data_validation(dv)
    dv.add("E3:E%d" % last)
    for k, fill in FILLS.items():
        ws3.conditional_formatting.add("E3:E%d" % last,
                                       CellIsRule(operator="equal", formula=['"%s"' % k], fill=fill))
    ws3.freeze_panes = "C3"
    ws3.auto_filter.ref = "A2:G%d" % last

    # ---- 조치 필요 ----
    ws4 = wb.create_sheet("조치 필요 항목")
    ws4.sheet_view.showGridLines = False
    for i, w in enumerate([4, 8, 22, 32, 34, 46, 12, 24], start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4["B2"] = "조치 필요 항목"
    ws4["B2"].font = Font(name="맑은 고딕", size=14, bold=True, color=NAVY)
    _hdr(ws4, 4, 2, ["No", "구분", "점검 항목", "기대값", "현재 상태", "판정", "조치 방안"])

    r4 = 5
    for it in checks:
        res = results.get(it["no"], {})
        v = res.get("verdict", "")
        if v not in (NOT_APPLIED, CHECK):
            continue
        out = res.get("out", "")
        if len(out) > 350:
            out = out[:350] + "…"
        for i, val in enumerate([it["no"], it["sec"], it["item"], it["exp"], out, v, res.get("note", "")], start=2):
            c = ws4.cell(r4, i, val)
            c.border = BORDER
            c.font = MONO if i == 6 else SMALL
            c.alignment = CENTER if i in (2, 7) else TOP
        ws4.cell(r4, 7).fill = FILLS.get(v, NA_FILL)
        ws4.row_dimensions[r4].height = min(110, max(26, 13 * min(out.count("\n") + 1, 7)))
        r4 += 1
    if r4 == 5:
        ws4.cell(5, 2, "조치 필요 항목 없음 — 전 항목 정상 적용").font = BODY
    else:
        ws4.auto_filter.ref = "B4:H%d" % (r4 - 1)
        ws4.freeze_panes = "B5"

    wb.save(outpath)
    return outpath, tot
